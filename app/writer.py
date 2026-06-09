"""Escritura incremental a archivos .xlsx (una hoja por archivo).

El servicio NO regenera el archivo de salida desde cero: solo agrega filas
nuevas. Si el archivo no existe, se crea con la fila de encabezado.

Decisión: usamos openpyxl directo (no pandas.to_excel) para no romper formato
ni hojas vecinas si el usuario ya tiene fórmulas o estilos en el archivo final
del KPI. La salida de este servicio es una tabla plana, sin fórmulas, para
auditarla y pegarla.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook, load_workbook

log = logging.getLogger(__name__)


def append_rows(
    path: str | Path,
    sheet_name: str,
    columns: Sequence[str],
    rows: Iterable[Sequence],
) -> int:
    """Agrega filas a `path` (hoja `sheet_name`). Crea archivo y hoja si no existen.

    - Si el archivo ya existe y la hoja también, valida que el encabezado
      coincida (mismas columnas en el mismo orden). Si no coincide, lanza error
      antes de tocar nada (mejor parar que escribir basura).
    - Devuelve la cantidad de filas escritas.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    rows = list(rows)

    if path.exists():
        wb = load_workbook(path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            existing_header = [c.value for c in ws[1]] if ws.max_row >= 1 else []
            if [h for h in existing_header if h is not None] != list(columns):
                raise ValueError(
                    f"El archivo {path} tiene un encabezado distinto en la hoja "
                    f"{sheet_name!r}. Esperado: {list(columns)}. "
                    f"Encontrado: {existing_header}. Movelo o borralo para regenerarlo."
                )
        else:
            ws = wb.create_sheet(sheet_name)
            ws.append(list(columns))
    else:
        wb = Workbook()
        # Workbook() crea una hoja por defecto llamada "Sheet": la reusamos
        # con el nombre que necesitamos.
        ws = wb.active
        ws.title = sheet_name
        ws.append(list(columns))

    for row in rows:
        ws.append(list(row))

    wb.save(path)
    log.info("writer: %d filas agregadas a %s (hoja %s)", len(rows), path, sheet_name)
    return len(rows)
