"""Parser del reporte diario de daños de Brasil (Furlong).

Adjunto: 'Reporte de Daños y equipos descargados DDMMYYYY.xlsx', hoja 'DAÑOS'.

Estructura particular de la hoja:
- Está segmentada en BLOQUES, uno por compañía (TRANSPORTE FURLONG, TTFA,
  AUTOPORT). Cada bloque empieza con una fila-encabezado:
      ['#', 'EQUIPO', 'COMPANIA', 'VIN', 'DAÑO', 'OBSERVACIONES']
- Bajo el encabezado vienen 12-20 filas plantilla numeradas; muchas vacías
  (sin EQUIPO/VIN/DAÑO).
- Cada bloque termina con una fila TOTAL: '#'=None, todas vacías salvo la
  columna 'DAÑO' que trae el conteo. Ese total NO se carga.

Una fila se considera "real" si trae al menos uno de EQUIPO, VIN o DAÑO con
contenido. La compañía se identifica por la columna COMPANIA dentro de cada
bloque (es más robusto que asumir el orden de los bloques).

Mapeo al esquema 'Brasil Impo DB' (corredor IMPO):
    EQUIPO          → Fleet N°
    VIN             → VIN
    DAÑO (texto)    → Cause
    OBSERVACIONES   → OBS.
    fecha reporte   → Fecha y Fecha de descarga
El resto de columnas quedan vacías (se completan a mano).

Deduplicación: Brasil no tiene un ID único confiable. Usamos doble red:
    1) processed_messages (un mismo mail no se procesa dos veces)
    2) row_key = sha1(fecha + VIN + texto del daño)[:16]
"""

from __future__ import annotations

import hashlib
import io
import logging
import re
from datetime import date, datetime
from typing import Iterable, List, Optional, Tuple

from openpyxl import load_workbook

log = logging.getLogger(__name__)

# Columnas destino — orden exacto de la hoja 'Brasil Impo DB' del KPI.
IMPO_COLUMNS: List[str] = [
    "LDR N°",
    "N°Caso",
    "Unidad",
    "Fecha",
    "Fecha de descarga",
    "Damage\ntype",
    "Damage\npart",
    "VIN",
    "Position",
    "Fleet N°",
    "Cause",
    "Responsible",
    "OBS.",
    "C/M",
    "C/M STATUS",
]

# Encabezado esperado dentro de cada bloque (en orden, las primeras 6).
_BLOCK_HEADER = ("#", "EQUIPO", "COMPANIA", "VIN", "DAÑO", "OBSERVACIONES")


class BrasilParseError(RuntimeError):
    """El adjunto no respeta el formato esperado."""


# ----------------------------------------------------------- fecha del archivo

_DATE_PATTERNS = [
    # 08-06-2026, 08_06_2026, 08/06/2026
    (re.compile(r"(?P<d>\d{2})[-_/](?P<m>\d{2})[-_/](?P<y>\d{4})"), "dmy"),
    # 2026-06-08
    (re.compile(r"(?P<y>\d{4})[-_/](?P<m>\d{2})[-_/](?P<d>\d{2})"), "ymd"),
    # 08062026
    (re.compile(r"(?<!\d)(?P<d>\d{2})(?P<m>\d{2})(?P<y>\d{4})(?!\d)"), "dmy"),
]


def date_from_filename(filename: str) -> Optional[date]:
    """Extrae la fecha del nombre del archivo. None si no encuentra."""
    for pattern, order in _DATE_PATTERNS:
        m = pattern.search(filename)
        if not m:
            continue
        d, mo, y = int(m["d"]), int(m["m"]), int(m["y"])
        try:
            return date(y, mo, d)
        except ValueError:
            continue
    return None


# ------------------------------------------------------------------- parsing

def _is_block_header(row: tuple) -> bool:
    """¿Esta fila es el encabezado `# / EQUIPO / COMPANIA / ...` de un bloque?"""
    if not row:
        return False
    cells = tuple(
        (str(c).strip().upper() if isinstance(c, str) else c) for c in row[: len(_BLOCK_HEADER)]
    )
    expected = tuple(h.upper() for h in _BLOCK_HEADER)
    return cells == expected


def _norm(value) -> str:
    """Normaliza un valor de celda a string limpio (vacío si nulo)."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _row_key(report_date: date, vin: str, dano: str) -> str:
    base = f"{report_date.isoformat()}|{vin.strip().upper()}|{dano.strip()}"
    return hashlib.sha1(base.encode("utf-8")).hexdigest()[:16]


def parse(
    file_bytes: bytes,
    report_date: date,
    company: str = "TTFA",
    sheet_name: str = "DAÑOS",
) -> List[dict]:
    """Lee la hoja DAÑOS y devuelve filas mapeadas al esquema IMPO para `company`.

    Cada dict tiene todas las columnas de IMPO_COLUMNS más un campo extra
    interno `__row_key__` para deduplicación.
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise BrasilParseError(
            f"El adjunto no tiene la hoja {sheet_name!r}. "
            f"Hojas encontradas: {wb.sheetnames}"
        )
    ws = wb[sheet_name]

    company_norm = company.strip().upper()
    out: List[dict] = []
    in_block = False
    current_company: Optional[str] = None

    for raw in ws.iter_rows(values_only=True):
        if _is_block_header(raw):
            in_block = True
            current_company = None  # se descubre con la primera fila numerada
            continue
        if not in_block:
            continue
        # raw: (#, EQUIPO, COMPANIA, VIN, DAÑO, OBSERVACIONES, ...)
        idx = raw[0] if len(raw) > 0 else None
        equipo = _norm(raw[1] if len(raw) > 1 else None)
        comp = _norm(raw[2] if len(raw) > 2 else None)
        vin = _norm(raw[3] if len(raw) > 3 else None)
        dano = _norm(raw[4] if len(raw) > 4 else None)
        obs = _norm(raw[5] if len(raw) > 5 else None)

        # Fila TOTAL del bloque: '#' vacío + casi todo vacío + DAÑO numérico.
        # La descartamos y cerramos el bloque hasta el próximo encabezado.
        if idx is None and not equipo and not comp and not vin:
            in_block = False
            current_company = None
            continue

        # Fija la compañía del bloque la primera vez que aparece.
        if current_company is None and comp:
            current_company = comp.upper()

        # Solo nos importa el bloque de la compañía configurada.
        if current_company != company_norm:
            continue

        # Filas plantilla vacías (sin EQUIPO, sin VIN, sin DAÑO real) se ignoran.
        if not equipo and not vin and not dano:
            continue

        record = {col: None for col in IMPO_COLUMNS}
        record["Fecha"] = report_date
        record["Fecha de descarga"] = report_date
        record["Fleet N°"] = equipo or None
        record["VIN"] = vin or None
        record["Cause"] = dano or None
        record["OBS."] = obs or None
        record["__row_key__"] = _row_key(report_date, vin, dano)
        out.append(record)

    log.info(
        "brasil: %d filas reales para compañía %s (fecha reporte %s)",
        len(out),
        company_norm,
        report_date.isoformat(),
    )
    return out


# -------------------------------------------------------------- deduplicación

def filter_new(records: Iterable[dict], known: Iterable[str]) -> List[dict]:
    """Filtra registros cuya row_key ya esté en `known`. Deduplica el lote también."""
    known_set = set(known)
    seen_in_batch: set[str] = set()
    out: List[dict] = []
    for rec in records:
        key = str(rec.get("__row_key__") or "")
        if not key or key in known_set or key in seen_in_batch:
            continue
        seen_in_batch.add(key)
        out.append(rec)
    return out


def records_to_rows(records: Iterable[dict]) -> List[list]:
    """Convierte registros a listas en el orden de IMPO_COLUMNS."""
    return [[rec.get(c) for c in IMPO_COLUMNS] for rec in records]


def keys_of(records: Iterable[dict]) -> List[str]:
    return [str(r.get("__row_key__")) for r in records if r.get("__row_key__")]


# ----------------------------------------------------------- helper para fecha

def fallback_date(received_iso: Optional[str]) -> date:
    """Convierte la fecha de recepción del mail (ISO) a date. Hoy como último recurso."""
    if received_iso:
        try:
            return datetime.fromisoformat(received_iso.replace("Z", "+00:00")).date()
        except ValueError:
            pass
    return datetime.now().date()


def resolve_report_date(filename: str, received_iso: Optional[str] = None) -> Tuple[date, str]:
    """Devuelve (fecha, fuente_del_dato) para auditar de dónde salió la fecha."""
    d = date_from_filename(filename)
    if d is not None:
        return d, "filename"
    return fallback_date(received_iso), "received"
