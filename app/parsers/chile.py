"""Parser del reporte de arribos de Chile (Furlong).

Hoja de origen: 'ARRIBOS' del adjunto 'REPORTE DE ARRIBOS.xlsx'.
Hoja destino:   'Pegar Chile' del libro KPI.

El archivo es ACUMULADO: cada envío trae todo lo cargado hasta el momento. El
servicio debe filtrar y dejar solo las filas cuya ID_Inspection no esté ya en
el estado.

Sobre la deduplicación:
- Un VIN con dos daños distintos aparece dos veces con dos ID_Inspection
  distintos → se conservan ambas.
- Filas idénticas (mismo VIN sin daños, etc.) comparten ID_Inspection y por
  eso se colapsan.
"""

from __future__ import annotations

import io
import logging
from typing import Iterable, List

from openpyxl import load_workbook

log = logging.getLogger(__name__)

# 19 columnas en el mismo orden que la hoja 'Pegar Chile' del libro KPI.
COLUMNS: List[str] = [
    "VIN",
    "ID_arribo",
    "ID_Inspection",
    "Modelo",
    "Posicion",
    "Fecha",
    "Equipo",
    "Lona",
    "Prosedencia",
    "Daño?",
    "Parte",
    "Tipo Daño",
    "Cuadrante",
    "Comentario",
    "Foto daño",
    "Validacion TCL",
    "Firma TCL",
    "Validacion transporte",
    "Firma transporte",
]

DEDUP_KEY = "ID_Inspection"


class ChileParseError(RuntimeError):
    """El adjunto no respeta el formato esperado (hoja u orden de columnas)."""


def parse(file_bytes: bytes, sheet_name: str = "ARRIBOS") -> List[dict]:
    """Lee el .xlsx de Chile y devuelve una lista de dicts (uno por fila)."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ChileParseError(
            f"El adjunto no tiene la hoja {sheet_name!r}. "
            f"Hojas encontradas: {wb.sheetnames}"
        )
    ws = wb[sheet_name]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        raise ChileParseError("La hoja está vacía (sin encabezado).")

    header_list = [str(h).strip() if h is not None else "" for h in header]
    # Tolerante: el archivo real puede traer alguna columna extra al final
    # (siempre y cuando las primeras 19 sean las esperadas en el mismo orden).
    if header_list[: len(COLUMNS)] != COLUMNS:
        raise ChileParseError(
            "Encabezado distinto al esperado en el reporte de Chile.\n"
            f"Esperado: {COLUMNS}\n"
            f"Encontrado: {header_list[: len(COLUMNS)]}"
        )

    out: List[dict] = []
    for raw in rows_iter:
        # Saltamos filas totalmente vacías (suele pasar al final).
        if raw is None or all(v is None or (isinstance(v, str) and not v.strip()) for v in raw):
            continue
        rec = {col: raw[i] if i < len(raw) else None for i, col in enumerate(COLUMNS)}
        # ID_Inspection es la clave; si llega vacía, descartamos la fila
        # (no podríamos deduplicar y no nos sirve).
        if not rec.get(DEDUP_KEY):
            continue
        out.append(rec)

    log.info("chile: %d filas leídas de la hoja %s", len(out), sheet_name)
    return out


def filter_new(records: Iterable[dict], known: Iterable[str]) -> List[dict]:
    """Devuelve solo las filas cuyo ID_Inspection no esté en `known`.

    También deduplica dentro del mismo lote: si el archivo trae dos filas con
    el mismo ID_Inspection, nos quedamos con la primera.
    """
    known_set = set(known)
    seen_in_batch: set[str] = set()
    out: List[dict] = []
    for rec in records:
        key = str(rec.get(DEDUP_KEY) or "").strip()
        if not key:
            continue
        if key in known_set or key in seen_in_batch:
            continue
        seen_in_batch.add(key)
        out.append(rec)
    return out


def records_to_rows(records: Iterable[dict]) -> List[list]:
    """Convierte registros a listas en el orden de COLUMNS, listas para append."""
    return [[rec.get(c) for c in COLUMNS] for rec in records]
